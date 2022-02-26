import datetime
import sys
import csv
import openpyxl
from pathlib import Path as P


def id_filter(row):
	if row[col_num].value in ids:
		return False
	else:
		return True


class Comparer():
	def main(self):
		print("Made by PunkOkami", "Published under GNU GPLv3 licence",
			  "Kod żródłowy: https://github.com/PunkOkami/xls_csv_comparer",
			  "Version: 1.2", "email adress: okami.github@gmail.com", "\n", sep="\n")
		print("---------------------------------------------------")
		csv_path = self.searcher("CSV", "Raport*.csv")
		xls_path = self.searcher("XLS", "eRej*[!-wynik].xlsx")
		csv_data = self.csv_analise(csv_path)
		xls_data = self.xls_analise(xls_path)
		csv_filtered = [row for row in csv_data["keep"] if row[1] not in xls_data["ids"]]
		xls_filtered = [row for row in xls_data["keep"] if row[0] not in csv_data["ids"]]
		print(len(csv_data["keep"]), len(xls_data["keep"]), sep="-----")
		error_rows = [row for row in csv_data["keep"] if row[0] == "" or row[3] != ""]
		print(f"Liczba rzędów z wartością 'Zrealizowana': {len(xls_data['keep'])}",
			  f"Liczba rzędów w pliku xls, która różni się od pliku csv: {len(xls_filtered)}",
			  f"Liczba rzędów w pliku csv, która różni się od pliku xls: {len(csv_filtered)}",
			  f"Liczba rzędów z możliwym błedem: {len(error_rows)}", sep="\n")
		result_xls_path = P(f"{xls_path.stem}-wynik{xls_path.suffix}")
		result_xls = openpyxl.Workbook()
		result_sheet = result_xls.active
		result_sheet.append([f"Liczba przeanalizowanych wpisów raportu z eRejestacji (plik XLSX): {len(xls_data['keep'])}"])
		result_sheet.append([f"Z czego {len(xls_filtered)} nie ma w raporcie z Gabinet.gov.pl:"])
		result_sheet.append(["Data i godzina rozpoczęcia terminu", "Wartość identyfikatora pacjenta"])
		for row in xls_filtered:
			result_sheet.append(row)
		result_sheet.append(["", "", ""])
		result_sheet.append([f"Liczba przeanalizowanych wpisów raportu z Gabinet.gov.pl (plik CSV): {len(csv_data['keep'])}"])
		result_sheet.append([f"Z czego {len(csv_filtered)} nie ma w raporcie z eRejestracji:"])
		result_sheet.append(["pm_root", "pacjent_ext", "pj_data_zapisania_baza_danych", "reg_aktywne"])
		for row in csv_filtered:
			result_sheet.append(row)
		result_sheet.append(["", "", ""])
		result_sheet.append([f"Liczba wpisów, które zawierają informację o potencjalnym błędzie: {len(error_rows)}"])
		for row in error_rows:
			result_sheet.append(row)
		result_xls.save(result_xls_path)
		print("----------------------------------------------")
		inn = input("Naciśnij Enter by zakończyć program")
		
	def searcher(self, filetype: str, name_regex: str) -> P:
		matching_files_list = list(P(P.cwd()).rglob(name_regex))
		if len(matching_files_list) != 1:
			if len(matching_files_list) == 0:
				inn = input(f"Program nie znalazł żadnego pliku {filetype}, naciśnij ENTER by zamknąć program\n>>>")
				sys.exit()
			else:
				status = True
				while status:
					print(
						f"Program znalazł więcej niż jeden plik {filetype}. Co program ma zrobić? Wpisz numer opcji, którą wybierasz",
						"1.Zamknij program, sam uporządkuję zawartość katalogu", "2.Weź najnowszy plik do analizy",
						"3.Pokaż mi listę plików do wyboru", sep="\n")
					option = input(">>>")
					if int(option) == 1:
						sys.exit()
					elif int(option) == 2:
						mtimes = {file.stat().st_mtime: file for file in matching_files_list}
						newest = max(mtimes.keys())
						file_path = mtimes[newest]
						status = False
					elif int(option) == 3:
						print(f"Pliki {filetype} znajdujące się w katalogu")
						for file in matching_files_list:
							mtime = file.stat().st_mtime
							print(f"{matching_files_list.index(file) + 1}. {file.name} --- {datetime.datetime.fromtimestamp(mtime)}")
						fnum = int(input(f"Wpisz numer pliku {filetype} jaki program ma wybrać\n>>>"))
						file_path = matching_files_list[fnum - 1]
						status = False
					else:
						print("Opcja niepoprawna, spróbuj ponownie\n---------------------------------------")
		else:
			file_path = matching_files_list[0]
		print("----------------------------------------------")
		print(f"Plik {filetype} wzięty do analizy to {file_path.name}")
		print("----------------------------------------------")
		return file_path
	
	def csv_analise(self, path):
		fcsv = open(path, newline="", encoding="utf-8")
		rcsv = csv.reader(fcsv, delimiter=";")
		rows = [row for row in rcsv]
		name_row = rows.pop(0)
		names = ["pm_root", "pacjent_ext", "pj_data_zapisania_baza_danych", "reg_aktywne"]
		error_col = name_row.index(names[0])
		id_col = name_row.index(names[1])
		time_col = name_row.index(names[2])
		reg_col = name_row.index(names[3])
		ids = set([row[id_col] for row in rows])
		keep = [(row[error_col], row[id_col], row[time_col], row[reg_col]) for row in rows]
		return {"keep": keep, "ids": ids}
	
	def xls_analise(self, path):
		fxls = openpyxl.load_workbook(path)
		sheet = fxls["Terminy i wizyty"]
		rows = [list(row) for row in sheet.iter_rows()]
		keep = [row for row in rows if row[3].value == "Zrealizowana"]
		id_row = [cell.value for cell in rows[0]]
		names = ["Data i godzina rozpoczęcia terminu", "Wartość identyfikatora pacjenta"]
		time_col = id_row.index(names[0])
		id_col = id_row.index(names[1])
		ids = set([row[id_col].value for row in keep])
		keep = [(row[id_col].value, row[time_col].value) for row in keep]
		return {"keep": keep, "ids":ids}

App = Comparer()
App.main()
